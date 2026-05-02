#!/usr/bin/env python3
"""Calibrate presentation-analysis scorer against the 18-clip ground truth.

Reads ``New Video Analysis_with comments.xlsx`` from the repo root, finds each
matching video file, extracts pose features once (cached to
``scripts/.presentation_features_cache.json``), then runs
:func:`src.presentation_scorer.score_presentation` and prints a per-clip /
per-category comparison plus accuracy.

Typical use
-----------

    # 1) Extract features (slow, do once):
    python scripts/calibrate_presentation_analysis.py --extract

    # 2) Score + compare (fast, iterate on thresholds with env vars):
    python scripts/calibrate_presentation_analysis.py --score

    # 3) Extract + score in one shot:
    python scripts/calibrate_presentation_analysis.py --all

Threshold tuning example (no redeploy needed)
---------------------------------------------

    PRES_ENG_GESTURE_HIGH=0.18 PRES_CONF_SWAY_HIGH=0.060 \
        python scripts/calibrate_presentation_analysis.py --score
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from typing import Any, Dict, List, Optional, Tuple

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(HERE)
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from src.presentation_scorer import score_presentation  # noqa: E402

GROUND_TRUTH_XLSX = os.path.join(REPO_ROOT, "New Video Analysis_with comments.xlsx")
GROUND_TRUTH_XLSX_10TYPES = os.path.join(REPO_ROOT, "Movement Combinations for Report 10 types.xlsx")
# Optional JSON sidecar for additional labelled clips (names + H/M/L per category).
# Keeps the original XLSX immutable while letting us extend the training set
# in-repo without touching openpyxl. See README in the file itself.
GROUND_TRUTH_EXTRA_JSON = os.path.join(REPO_ROOT, "extra_ground_truth.json")
FEATURE_CACHE = os.path.join(HERE, ".presentation_features_cache.json")

CATEGORIES = [
    "eye_contact",
    "uprightness",
    "stance",
    "engaging",
    "adaptability",
    "confidence",
    "authority",
]

# Map Excel-column label ↔ scorer field.
XLSX_COLS = [
    ("Eye Contact", "eye_contact"),
    ("Uprightness", "uprightness"),
    ("Stance", "stance"),
    ("Engagement", "engaging"),
    ("Adaptability", "adaptability"),
    ("Confidence", "confidence"),
    ("Authority", "authority"),
]


def _slug(name: str) -> str:
    s = re.sub(r"\s+", "", str(name or "").lower())
    return s


def _norm_band(val: Any) -> Optional[str]:
    s = str(val or "").strip().upper()
    if s in ("H", "HIGH"):
        return "High"
    if s in ("M", "MODERATE", "MOD"):
        return "Moderate"
    if s in ("L", "LOW"):
        return "Low"
    return None


def load_ground_truth() -> List[Dict[str, Any]]:
    """Parse both ground-truth Excel files into a merged record list.

    Sources:
      - ``New Video Analysis_with comments.xlsx`` — the original 18-clip set
        (Andre / Aom / Meiji / ...). Free-form Name column. Labels use H/M/L.
      - ``Movement Combinations for Report 10 types.xlsx`` Sheet2 —
        10 archetypes keyed by ``Type N`` with full-word bands (High/Moderate/Low).

    Records are identified by a non-empty Name cell. Later sources can override
    earlier ones by slug, but in practice the Type 1..10 names don't clash with
    the original set.
    """
    records: List[Dict[str, Any]] = []
    seen_slugs: set = set()

    # --- Source 1: original 18-clip spreadsheet ---------------------------
    if os.path.exists(GROUND_TRUTH_XLSX):
        records.extend(_load_primary_xlsx())
        seen_slugs.update(r["slug"] for r in records)

    # --- Source 2: 10-types archetype spreadsheet -------------------------
    if os.path.exists(GROUND_TRUTH_XLSX_10TYPES):
        for rec in _load_10types_xlsx():
            if rec["slug"] in seen_slugs:
                continue
            records.append(rec)
            seen_slugs.add(rec["slug"])

    # --- Source 3: optional JSON sidecar (extra labelled clips) -----------
    # Lets us add new students/professor labels without touching the XLSX.
    # JSON shape: list of objects, each:
    #   {"name": "Tiffany 2", "categories": {"eye_contact": "High",
    #    "uprightness": "Moderate", "stance": "Low", ...}}
    if os.path.exists(GROUND_TRUTH_EXTRA_JSON):
        for rec in _load_extra_json():
            if rec["slug"] in seen_slugs:
                continue
            records.append(rec)
            seen_slugs.add(rec["slug"])

    return records


def _load_extra_json() -> List[Dict[str, Any]]:
    """Parse ``extra_ground_truth.json`` if present.

    Each entry should already be a dict with at minimum a ``name`` and a
    ``categories`` mapping. Bands are normalised through :func:`_norm_band` so
    callers can use H/M/L shorthand or the full word.
    """
    try:
        with open(GROUND_TRUTH_EXTRA_JSON, "r", encoding="utf-8") as fh:
            raw = json.load(fh)
    except Exception as exc:  # pragma: no cover — operator-facing tooling
        print(f"[extra-gt] could not load {GROUND_TRUTH_EXTRA_JSON}: {exc}")
        return []
    if not isinstance(raw, list):
        print(f"[extra-gt] expected a list at top level, got {type(raw).__name__}")
        return []
    out: List[Dict[str, Any]] = []
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip()
        cats_raw = entry.get("categories") or {}
        if not name or not isinstance(cats_raw, dict):
            continue
        cats: Dict[str, str] = {}
        for key, val in cats_raw.items():
            band = _norm_band(val)
            if band and key in CATEGORIES:
                cats[key] = band
        if not cats:
            continue
        out.append({
            "name": name,
            "slug": _slug(name),
            "categories": cats,
            "source": "extra",
        })
    return out


def _load_primary_xlsx() -> List[Dict[str, Any]]:
    import openpyxl  # local import so the rest of the script runs without it

    wb = openpyxl.load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb.active
    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()
        if name_str.lower() == "name":
            continue  # header
        cats: Dict[str, str] = {}
        # Columns: 0=Name, 1..7=categories in XLSX_COLS order.
        for i, (_, key) in enumerate(XLSX_COLS, start=1):
            band = _norm_band(row[i] if i < len(row) else None)
            if band:
                cats[key] = band
        if not cats:
            continue
        records.append({
            "name": name_str,
            "slug": _slug(name_str),
            "categories": cats,
            "source": "primary",
        })
    return records


# Column order in Sheet2 of the 10-types workbook.
#   Type | Name | Eye contact | Stance | Uprightness | Engaging | Authority | Confidence | Adaptability
_TYPES_COLS: List[Tuple[int, str]] = [
    (2, "eye_contact"),
    (3, "stance"),
    (4, "uprightness"),
    (5, "engaging"),
    (6, "authority"),
    (7, "confidence"),
    (8, "adaptability"),
]


def _load_10types_xlsx() -> List[Dict[str, Any]]:
    """Parse Sheet2 of ``Movement Combinations for Report 10 types.xlsx``.

    Each row is one archetype keyed by Type 1..10. Record ``name`` is
    ``"Type N"`` so video matching lands on ``Type N.mov``.
    """
    import openpyxl

    wb = openpyxl.load_workbook(GROUND_TRUTH_XLSX_10TYPES, data_only=True)
    if "Sheet2" not in wb.sheetnames:
        return []
    ws = wb["Sheet2"]
    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        type_id = row[0]
        if type_id is None:
            continue
        try:
            n = int(type_id)
        except (TypeError, ValueError):
            continue
        name_str = f"Type {n}"
        cats: Dict[str, str] = {}
        for idx, key in _TYPES_COLS:
            band = _norm_band(row[idx] if idx < len(row) else None)
            if band:
                cats[key] = band
        if not cats:
            continue
        records.append({
            "name": name_str,
            "slug": _slug(name_str),
            "categories": cats,
            "source": "10types",
            "display_name": str(row[1]).strip() if row[1] else name_str,
        })
    return records


def find_video_for_record(rec: Dict[str, Any]) -> Optional[str]:
    """Match a ground-truth record to a video file in REPO_ROOT by slug.

    ``Lisa 2`` → ``Lisa2.mp4`` / ``Lisa1`` → ``Lisa.mp4`` if only Lisa.mp4 exists.
    """
    target_slug = rec["slug"]
    exts = (".mp4", ".mov", ".m4v", ".webm", ".avi", ".mkv")
    candidates = []
    for fn in os.listdir(REPO_ROOT):
        if not fn.lower().endswith(exts):
            continue
        stem = os.path.splitext(fn)[0]
        stem_slug = _slug(stem)
        if stem_slug == target_slug:
            return os.path.join(REPO_ROOT, fn)
        candidates.append((stem_slug, fn))

    # Fallback: "Lisa1" ↔ "Lisa"
    if target_slug.endswith("1"):
        bare = target_slug[:-1]
        for ss, fn in candidates:
            if ss == bare:
                return os.path.join(REPO_ROOT, fn)

    return None


def extract_features(
    records: List[Dict[str, Any]],
    cache_path: str = FEATURE_CACHE,
) -> Dict[str, Dict[str, Any]]:
    """Run MediaPipe on each video once; cache features for fast re-scoring."""
    cache: Dict[str, Dict[str, Any]] = {}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as fh:
                cache = json.load(fh)
        except Exception:
            cache = {}

    # Heavy imports only when we actually need to extract.
    from src.report_core import (
        analyze_first_impression_from_video,
        analyze_video_mediapipe,
    )

    for rec in records:
        name = rec["name"]
        if name in cache:
            print(f"[cache] {name} — already extracted")
            continue

        video_path = find_video_for_record(rec)
        if not video_path:
            print(f"[miss] {name} — no matching video file")
            continue

        print(f"[extract] {name} ← {os.path.basename(video_path)}")
        try:
            fi = analyze_first_impression_from_video(video_path, audience_mode="one")
            result = analyze_video_mediapipe(
                video_path,
                sample_fps=3,
                max_frames=200,
            )
        except Exception as e:
            print(f"  !! failed: {e}")
            continue

        cache[name] = {
            "video_file": os.path.basename(video_path),
            "first_impression": [fi.eye_contact_pct, fi.upright_pct, fi.stance_stability],
            "category_features": result.get("category_features", {}),
            "effort_counts": result.get("effort_counts", {}),
            "analyzed_frames": result.get("analyzed_frames", 0),
        }
        # Save after each clip so we don't lose progress if MediaPipe crashes.
        with open(cache_path, "w", encoding="utf-8") as fh:
            json.dump(cache, fh, indent=2)

    return cache


def score_and_compare(
    records: List[Dict[str, Any]],
    cache: Dict[str, Dict[str, Any]],
    verbose: bool = True,
) -> Dict[str, Any]:
    """Score each cached clip and compare to Excel ground truth.

    Returns a summary dict with per-category accuracy and a per-clip row
    table ready to print.
    """
    rows: List[Dict[str, Any]] = []
    totals = {cat: {"hit": 0, "total": 0, "adj": 0} for cat in CATEGORIES}

    for rec in records:
        name = rec["name"]
        feats = cache.get(name)
        if not feats:
            if verbose:
                print(f"[skip] {name} — no cached features")
            continue
        truth = rec["categories"]

        pred = score_presentation(
            first_impression_pct=tuple(feats["first_impression"]),
            category_features=feats.get("category_features", {}),
            analysis_result={
                "analyzed_frames": feats.get("analyzed_frames", 0),
                "effort_counts": feats.get("effort_counts", {}),
            },
        )

        row = {"name": name}
        for cat in CATEGORIES:
            t = truth.get(cat)
            p = pred.get(cat)
            row[f"truth_{cat}"] = t
            row[f"pred_{cat}"] = p
            if t:
                totals[cat]["total"] += 1
                if t == p:
                    totals[cat]["hit"] += 1
                # "Adjacent" counts M↔H and M↔L as near-misses.
                if _is_adjacent(t, p):
                    totals[cat]["adj"] += 1
        row["rationale"] = pred.get("rationale", {})
        row["overview"] = pred.get("overview")
        rows.append(row)

    if verbose:
        _print_table(rows)
        _print_accuracy(totals)

    return {"rows": rows, "totals": totals}


def _is_adjacent(t: str, p: str) -> bool:
    ladder = {"Low": 0, "Moderate": 1, "High": 2}
    return (t in ladder and p in ladder) and abs(ladder[t] - ladder[p]) <= 1


def _print_table(rows: List[Dict[str, Any]]) -> None:
    cats = CATEGORIES
    col_w = 4
    header_cats = " | ".join(f"{_short(c):>{col_w * 2 + 3}}" for c in cats)
    print()
    print("Per-clip predictions vs ground truth (T=truth, P=pred, ✓/✗/~):")
    print(f"{'Name':<14} | {header_cats}")
    print("-" * (15 + len(cats) * (col_w * 2 + 6)))
    for row in rows:
        cells = []
        for cat in cats:
            t = row.get(f"truth_{cat}") or "-"
            p = row.get(f"pred_{cat}") or "-"
            if t == p:
                flag = "✓"
            elif t != "-" and p != "-" and _is_adjacent(t, p):
                flag = "~"
            else:
                flag = "✗"
            cells.append(f"{_letter(t)}/{_letter(p)} {flag}")
        print(f"{row['name']:<14} | " + " | ".join(f"{c:>{col_w * 2 + 3}}" for c in cells))


def _print_accuracy(totals: Dict[str, Dict[str, int]]) -> None:
    print()
    print("Per-category accuracy:")
    print(f"{'Category':<14} | {'Exact':>8} | {'Adj(±1)':>8}")
    print("-" * 40)
    overall_hit = overall_adj = overall_n = 0
    for cat in CATEGORIES:
        t = totals[cat]
        n = t["total"]
        h = t["hit"]
        a = t["adj"]
        overall_hit += h
        overall_adj += a
        overall_n += n
        pct_h = (h / n * 100) if n else 0.0
        pct_a = (a / n * 100) if n else 0.0
        print(f"{cat:<14} | {h}/{n} {pct_h:5.1f}% | {a}/{n} {pct_a:5.1f}%")
    if overall_n:
        print("-" * 40)
        print(
            f"{'TOTAL':<14} | "
            f"{overall_hit}/{overall_n} {overall_hit / overall_n * 100:5.1f}% | "
            f"{overall_adj}/{overall_n} {overall_adj / overall_n * 100:5.1f}%"
        )


def _short(cat: str) -> str:
    mapping = {
        "eye_contact": "Eye",
        "uprightness": "Upr",
        "stance": "Stn",
        "engaging": "Eng",
        "adaptability": "Adp",
        "confidence": "Cnf",
        "authority": "Aut",
    }
    return mapping.get(cat, cat[:3])


def _letter(band: str) -> str:
    if band == "High":
        return "H"
    if band == "Moderate":
        return "M"
    if band == "Low":
        return "L"
    return "-"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--extract", action="store_true", help="Extract features (runs MediaPipe)")
    parser.add_argument("--score", action="store_true", help="Score cached features vs Excel truth")
    parser.add_argument("--all", action="store_true", help="Extract + score")
    parser.add_argument("--cache", default=FEATURE_CACHE)
    parser.add_argument("--verbose-reasons", action="store_true", help="Print per-clip reason lists")
    args = parser.parse_args()

    if not (args.extract or args.score or args.all):
        parser.print_help()
        return 1

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    if not os.path.exists(GROUND_TRUTH_XLSX):
        print(f"[error] Missing ground truth: {GROUND_TRUTH_XLSX}", file=sys.stderr)
        return 2

    records = load_ground_truth()
    print(f"[gt] {len(records)} ground-truth records loaded")

    cache: Dict[str, Dict[str, Any]] = {}
    if args.extract or args.all:
        cache = extract_features(records, cache_path=args.cache)
    else:
        if os.path.exists(args.cache):
            with open(args.cache, "r", encoding="utf-8") as fh:
                cache = json.load(fh)

    if args.score or args.all:
        summary = score_and_compare(records, cache, verbose=True)
        if args.verbose_reasons:
            print("\nRationale per clip:")
            for row in summary["rows"]:
                print(f"\n--- {row['name']} ---")
                for cat in CATEGORIES:
                    t = row.get(f"truth_{cat}")
                    p = row.get(f"pred_{cat}")
                    r = (row.get("rationale") or {}).get(cat, [])
                    flag = "✓" if t == p else ("~" if _is_adjacent(t or "", p or "") else "✗")
                    print(f"  {cat:<14} truth={t} pred={p} {flag} : {', '.join(r)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
